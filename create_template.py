# create_template.py
import openpyxl
from openpyxl.styles import Font, PatternFill

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Contacts"

# Headers - ADD 'groups' column
headers = ['phone', 'name', 'notes', 'labels', 'groups']  # ⬅️ UPDATED
ws.append(headers)

# Style headers
header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
header_font = Font(bold=True, color="000000")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Sample data - ADD groups column
sample_data = [
    ['919876543210', 'John Doe', 'VIP customer', 'customer,vip', 'sales,vip'],  # ⬅️ UPDATED
    ['919876543211', 'Jane Smith', 'Regular customer', 'customer', 'marketing'],  # ⬅️ UPDATED
    ['919876543212', 'Bob Johnson', 'Supplier', 'supplier,important', 'procurement,suppliers']  # ⬅️ UPDATED
]

for row in sample_data:
    ws.append(row)

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20  # ⬅️ ADD THIS

# Save
wb.save('templates/contacts_template.xlsx')
print("✅ Template created: templates/contacts_template.xlsx")